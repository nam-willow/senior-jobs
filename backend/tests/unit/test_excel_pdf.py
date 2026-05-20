"""
EX-* Excel/PDF 생성기 단위 테스트.
MinIO, DB 없이 순수 바이트 생성만 검증.
"""
import io

import pytest
from openpyxl import load_workbook

from app.utils.excel_generator import (
    generate_consultation_log_excel,
    generate_salary_statement_excel,
    generate_work_log_excel,
)
from app.utils.pdf_generator import (
    generate_consultation_log_pdf,
    generate_salary_statement_pdf,
)

# ── EX-01 ~ EX-05: Excel 생성 ────────────────────────────────────────────────

class TestWorkLogExcel:
    def _seniors(self, n=2):
        return [
            {
                "name": f"어르신{i}",
                "workplace": f"근무장소{i}",
                "rows": [
                    {"date": f"2026/05/{d:02d}", "hours": ""}
                    for d in range(1, 11)
                ],
            }
            for i in range(1, n + 1)
        ]

    def test_EX01_returns_bytes(self):
        """EX-01: generate_work_log_excel — 반환값이 bytes."""
        result = generate_work_log_excel(2026, 5, self._seniors())
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_EX02_sheet_per_senior(self):
        """EX-02: 어르신 수만큼 시트 생성."""
        seniors = self._seniors(3)
        result = generate_work_log_excel(2026, 5, seniors)
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 3

    def test_EX03_sheet_title_in_first_row(self):
        """EX-03: 첫 행에 연도 포함 제목."""
        result = generate_work_log_excel(2026, 5, self._seniors(1))
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert "2026" in str(ws["A1"].value)
        assert "근무일지" in str(ws["A1"].value)

    def test_EX04_empty_seniors(self):
        """EX-04: 어르신 없으면 유효한 Excel 반환 (안내 시트 1개)."""
        result = generate_work_log_excel(2026, 5, [])
        assert isinstance(result, bytes)
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 1  # "데이터 없음" 안내 시트


class TestSalaryStatementExcel:
    def _records(self):
        return [
            {"name": "김노인", "birth_date": "1950-01-01", "worked_hours": 30.0, "amount_paid": 300000},
            {"name": "이노인", "birth_date": "1948-05-12", "worked_hours": 25.5, "amount_paid": 255000},
        ]

    def test_EX05_salary_excel_bytes(self):
        """EX-05: generate_salary_statement_excel — 반환값이 bytes."""
        result = generate_salary_statement_excel(2026, 5, "공익활동 1단", self._records())
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_EX06_salary_has_total_row(self):
        """EX-06: 합계행 포함 (총 행 수 = 헤더+데이터+합계)."""
        records = self._records()
        result = generate_salary_statement_excel(2026, 5, "공익활동 1단", records)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # 행 2=헤더, 3~4=데이터, 5=합계 → 최소 5행
        assert ws.max_row >= 2 + len(records) + 1

    def test_EX07_consultation_excel_bytes(self):
        """EX-07: generate_consultation_log_excel — 반환값이 bytes."""
        logs = [
            {
                "consultation_date": "2026-05-01 10:00:00",
                "method": "phone",
                "content": "건강 상태 확인",
                "memo": "특이사항 없음",
                "social_worker_name": "홍길동",
                "default_session_hours": 3,
            }
        ]
        result = generate_consultation_log_excel(logs, senior_name="김노인")
        assert isinstance(result, bytes)
        assert len(result) > 0


# ── EX-08 ~ EX-10: PDF 생성 ─────────────────────────────────────────────────

class TestPdfGenerator:
    def test_EX08_salary_pdf_bytes(self):
        """EX-08: generate_salary_statement_pdf — bytes, PDF 매직 바이트."""
        records = [
            {"name": "김노인", "birth_date": "1950-01-01", "worked_hours": 30.0, "amount_paid": 300000},
        ]
        result = generate_salary_statement_pdf(2026, 5, "공익활동 1단", records)
        assert isinstance(result, bytes)
        assert result[:4] == b"%PDF"

    def test_EX09_consultation_pdf_bytes(self):
        """EX-09: generate_consultation_log_pdf — bytes, PDF 매직 바이트."""
        logs = [
            {
                "consultation_date": "2026-05-01",
                "method": "visit",
                "content": "방문 상담 진행",
                "memo": "",
                "social_worker_name": "홍길동",
                "default_session_hours": 3,
            }
        ]
        result = generate_consultation_log_pdf(logs, senior_name="김노인")
        assert isinstance(result, bytes)
        assert result[:4] == b"%PDF"

    def test_EX10_salary_pdf_empty_records(self):
        """EX-10: 레코드 없어도 PDF 생성 가능 (빈 테이블)."""
        result = generate_salary_statement_pdf(2026, 5, "테스트", [])
        assert isinstance(result, bytes)
        assert len(result) > 0
