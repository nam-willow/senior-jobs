"""
Excel 생성 유틸리티.

근무일지: 6열 구성 (날짜/성명/근무장소/일한시간/담당자(인)/사회복지사(인))
급여대장: APPROVED 상태 기준, 합계행 포함
상담일지: 단건/다건
"""
import io
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 공통 스타일 헬퍼 ──────────────────────────────────────────────────────────

_THIN = Side(style="thin")
_THICK = Side(style="medium")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_THICK = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_FILL = PatternFill("solid", fgColor="DAEEF3")


def _apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _BORDER_ALL


# ── 근무일지 (WorkLog) ────────────────────────────────────────────────────────

def generate_work_log_excel(
    year: int,
    month: int,
    seniors: List[Dict[str, Any]],
) -> bytes:
    """
    근무일지 Excel 생성.

    seniors: [
      {
        "name": str,
        "workplace": str,
        "rows": [{"date": str, "hours": float}, ...]
      }
    ]
    어르신 1인 = 1시트
    """
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    if not seniors:
        ws = wb.create_sheet(title="데이터 없음")
        ws["A1"] = "출력할 어르신 데이터가 없습니다."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    title = f"{year}년도 노인일자리 근무일지"

    # 열 너비 비율: 날짜=1 / 성명=1 / 근무장소=4 / 일한시간=1 / 담당자=1 / 사회복지사=1
    col_widths = [10, 10, 40, 10, 12, 14]
    headers = ["날짜", "성명", "근무장소", "일한시간", "담당자 (인)", "사회복지사 (인)"]

    for idx, senior in enumerate(seniors, start=1):
        sheet_name = senior["name"][:31]  # 시트명 31자 제한
        ws = wb.create_sheet(title=sheet_name)

        # 제목 행
        ws.merge_cells("A1:F1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = _CENTER

        # 어르신명 + 근무장소 행
        ws.merge_cells("A2:B2")
        ws["A2"] = f"성명: {senior['name']}"
        ws["A2"].alignment = _CENTER
        ws.merge_cells("C2:F2")
        ws["C2"] = f"근무장소: {senior.get('workplace', '')}"
        ws["C2"].alignment = _CENTER

        # 헤더 행
        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = _CENTER
            cell.fill = _HEADER_FILL
            ws.column_dimensions[get_column_letter(col)].width = width

        # 데이터 행
        for row_idx, row_data in enumerate(senior.get("rows", []), start=4):
            date_val = row_data.get("date", "")
            ws.cell(row=row_idx, column=1, value=date_val).alignment = _CENTER
            ws.cell(row=row_idx, column=2, value=senior["name"]).alignment = _CENTER
            ws.cell(row=row_idx, column=3, value=senior.get("workplace", "")).alignment = _CENTER
            ws.cell(row=row_idx, column=4, value=row_data.get("hours", "")).alignment = _CENTER
            ws.cell(row=row_idx, column=5, value="").alignment = _CENTER  # 담당자 서명란
            ws.cell(row=row_idx, column=6, value="").alignment = _CENTER  # 복지사 서명란

        last_data_row = 3 + len(senior.get("rows", []))
        _apply_border(ws, 3, last_data_row, 1, 6)

        # 인쇄 설정 (A4 가로)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9  # A4
        ws.print_area = f"A1:F{last_data_row}"
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 급여대장 (Salary Statement) ───────────────────────────────────────────────

def generate_salary_statement_excel(
    year: int,
    month: int,
    business_unit_name: str,
    records: List[Dict[str, Any]],
) -> bytes:
    """
    급여대장 Excel 생성.

    records: APPROVED 상태 근무기록
    [
      {
        "name": str,
        "birth_date": str,
        "worked_hours": float,
        "amount_paid": int,
      }
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "급여대장"

    title = f"{year}년 {month}월 급여대장 — {business_unit_name}"
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = _CENTER

    headers = ["사업단명", "연월", "성명", "생년월일", "근무시간", "지급금액", "서명"]
    col_widths = [20, 12, 12, 14, 12, 16, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = w

    total_amount = 0
    for row_idx, rec in enumerate(records, start=3):
        ws.cell(row=row_idx, column=1, value=business_unit_name).alignment = _CENTER
        ws.cell(row=row_idx, column=2, value=f"{year}-{month:02d}").alignment = _CENTER
        ws.cell(row=row_idx, column=3, value=rec["name"]).alignment = _CENTER
        ws.cell(row=row_idx, column=4, value=rec.get("birth_date", "")).alignment = _CENTER
        ws.cell(row=row_idx, column=5, value=rec.get("worked_hours", 0)).alignment = _CENTER
        ws.cell(row=row_idx, column=6, value=rec.get("amount_paid", 0)).alignment = _CENTER
        ws.cell(row=row_idx, column=7, value="").alignment = _CENTER  # 서명란
        total_amount += rec.get("amount_paid", 0)

    # 합계 행
    sum_row = 3 + len(records)
    ws.merge_cells(f"A{sum_row}:E{sum_row}")
    ws.cell(row=sum_row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=sum_row, column=1).alignment = _CENTER
    ws.cell(row=sum_row, column=6, value=total_amount).font = Font(bold=True)
    ws.cell(row=sum_row, column=6).alignment = _CENTER

    _apply_border(ws, 2, sum_row, 1, 7)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 상담일지 (Consultation Log) ───────────────────────────────────────────────

def generate_consultation_log_excel(
    logs: List[Dict[str, Any]],
    senior_name: str = "",
) -> bytes:
    """
    상담일지 Excel 생성.

    logs: [
      {
        "consultation_date": str,
        "method": str,
        "content": str,
        "memo": str,
        "social_worker_name": str,
        "default_session_hours": int,
      }
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "상담일지"

    title = f"상담일지{' — ' + senior_name if senior_name else ''}"
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = _CENTER

    headers = ["상담일시", "상담방법", "상담내용", "담당자", "기본시간", "메모"]
    col_widths = [18, 12, 40, 14, 10, 20]
    method_map = {"phone": "전화", "visit": "방문", "in_person": "내방", "other": "기타"}

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = w

    for row_idx, log in enumerate(logs, start=3):
        ws.cell(row=row_idx, column=1, value=log.get("consultation_date", "")).alignment = _CENTER
        ws.cell(row=row_idx, column=2, value=method_map.get(log.get("method", ""), log.get("method", ""))).alignment = _CENTER
        ws.cell(row=row_idx, column=3, value=log.get("content", "")).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=4, value=log.get("social_worker_name", "")).alignment = _CENTER
        ws.cell(row=row_idx, column=5, value=log.get("default_session_hours", "")).alignment = _CENTER
        ws.cell(row=row_idx, column=6, value=log.get("memo", "")).alignment = Alignment(wrap_text=True, vertical="top")

    last_row = 2 + len(logs)
    _apply_border(ws, 2, last_row, 1, 6)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
